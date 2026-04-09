import os
from datetime import datetime
from zoneinfo import ZoneInfo

import discord
from discord.ext import commands, tasks
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook

load_dotenv()

TOKEN = os.getenv("DISCORD_TOKEN")
CHANNEL_ID = int(os.getenv("CHANNEL_ID"))
TZ = ZoneInfo(os.getenv("TZ", "Asia/Karachi"))

MORNING_HOUR = int(os.getenv("MORNING_HOUR", "10"))
MORNING_MINUTE = int(os.getenv("MORNING_MINUTE", "0"))
EVENING_HOUR = int(os.getenv("EVENING_HOUR", "18"))
EVENING_MINUTE = int(os.getenv("EVENING_MINUTE", "45"))

EXCEL_FILE = "daily_updates.xlsx"

intents = discord.Intents.default()
intents.message_content = True
intents.members = True
intents.guilds = True

bot = commands.Bot(command_prefix="!", intents=intents)

session_state = {
    "date": None,
    "morning": {
        "active": False,
        "status_message_id": None,
        "replied_users": set(),
        "user_steps": {}
    },
    "evening": {
        "active": False,
        "status_message_id": None,
        "replied_users": set(),
        "user_steps": {}
    },
}


def ensure_workbook():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Updates"
        ws.append([
            "Date",
            "User ID",
            "Username",
            "Display Name",
            "Morning Working On",
            "Morning Blockers / Notes",
            "Morning Timestamp",
            "Evening Achieved",
            "Evening Pending / Blockers",
            "Evening Timestamp",
        ])
        wb.save(EXCEL_FILE)


def save_update(date_str, user, update_type, part1, part2, timestamp_str):
    ensure_workbook()
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Updates"]

    target_row = None

    for row in range(2, ws.max_row + 1):
        row_date = ws.cell(row=row, column=1).value
        row_user_id = ws.cell(row=row, column=2).value
        if str(row_date) == date_str and str(row_user_id) == str(user.id):
            target_row = row
            break

    if target_row is None:
        target_row = ws.max_row + 1
        ws.cell(row=target_row, column=1, value=date_str)
        ws.cell(row=target_row, column=2, value=str(user.id))
        ws.cell(row=target_row, column=3, value=str(user))
        ws.cell(row=target_row, column=4, value=message_safe_display_name(user))

    if update_type == "morning":
        ws.cell(row=target_row, column=5, value=part1)
        ws.cell(row=target_row, column=6, value=part2)
        ws.cell(row=target_row, column=7, value=timestamp_str)
    elif update_type == "evening":
        ws.cell(row=target_row, column=8, value=part1)
        ws.cell(row=target_row, column=9, value=part2)
        ws.cell(row=target_row, column=10, value=timestamp_str)

    wb.save(EXCEL_FILE)


def message_safe_display_name(user):
    return getattr(user, "display_name", str(user))


def get_now():
    return datetime.now(TZ)


def get_today_str():
    return get_now().strftime("%Y-%m-%d")


def get_target_members(guild):
    return [member for member in guild.members if not member.bot]


def reset_for_new_day():
    today = get_today_str()
    if session_state["date"] != today:
        session_state["date"] = today
        session_state["morning"] = {
            "active": False,
            "status_message_id": None,
            "replied_users": set(),
            "user_steps": {}
        }
        session_state["evening"] = {
            "active": False,
            "status_message_id": None,
            "replied_users": set(),
            "user_steps": {}
        }


async def update_status_message(channel, guild, session_type):
    state = session_state[session_type]
    status_message_id = state["status_message_id"]

    if not status_message_id:
        return

    try:
        status_message = await channel.fetch_message(status_message_id)
    except discord.NotFound:
        return

    total_users = len(get_target_members(guild))
    replied = len(state["replied_users"])
    pending = max(total_users - replied, 0)

    label = "Morning" if session_type == "morning" else "Evening"

    content = (
        f"**{label} Status**\n"
        f"Replied: **{replied}/{total_users}**\n"
        f"Pending: **{pending}**"
    )

    await status_message.edit(content=content)


async def start_session(session_type):
    reset_for_new_day()

    channel = bot.get_channel(CHANNEL_ID)
    if channel is None:
        print("Channel not found. Check CHANNEL_ID.")
        return

    guild = channel.guild
    if guild is None:
        print("Guild not found.")
        return

    state = session_state[session_type]
    state["active"] = True
    state["replied_users"].clear()
    state["user_steps"].clear()

    if session_type == "morning":
        prompt_text = "🌞 **Morning Update Time**\nReply when ready."
    else:
        prompt_text = "🌙 **Evening Update Time**\nReply when ready."

    await channel.send(prompt_text)
    status_message = await channel.send("Preparing status...")

    state["status_message_id"] = status_message.id
    await update_status_message(channel, guild, session_type)


@bot.event
async def on_ready():
    print(f"Logged in as {bot.user}")
    ensure_workbook()

    for guild in bot.guilds:
        try:
            await guild.chunk()
        except Exception as e:
            print(f"Could not chunk guild {guild.name}: {e}")

    if not scheduler.is_running():
        scheduler.start()


@tasks.loop(minutes=1)
async def scheduler():
    reset_for_new_day()

    now = get_now()

    if now.hour == MORNING_HOUR and now.minute == MORNING_MINUTE:
        if not session_state["morning"]["active"]:
            await start_session("morning")

    if now.hour == EVENING_HOUR and now.minute == EVENING_MINUTE:
        if not session_state["evening"]["active"]:
            await start_session("evening")


@bot.command()
async def testmorning(ctx):
    await start_session("morning")
    await ctx.send("Morning session started.")


@bot.command()
async def testevening(ctx):
    await start_session("evening")
    await ctx.send("Evening session started.")


@bot.command()
async def closemorning(ctx):
    session_state["morning"]["active"] = False
    await ctx.send("Morning session closed.")


@bot.command()
async def closeevening(ctx):
    session_state["evening"]["active"] = False
    await ctx.send("Evening session closed.")


@bot.event
async def on_message(message):
    if message.author.bot:
        return

    await bot.process_commands(message)

    if message.channel.id != CHANNEL_ID:
        return

    reset_for_new_day()

    guild = message.guild
    if guild is None:
        return

    user_id = message.author.id
    today_str = get_today_str()
    now_str = get_now().strftime("%Y-%m-%d %H:%M:%S")

    # MORNING FLOW
    if session_state["morning"]["active"] and user_id not in session_state["morning"]["replied_users"]:
        user_steps = session_state["morning"]["user_steps"]

        if user_id not in user_steps:
            user_steps[user_id] = {
                "step": 1,
                "part1": "",
                "part2": ""
            }
            await message.reply("Working on")
            return

        elif user_steps[user_id]["step"] == 1 and user_steps[user_id]["part1"] == "":
            user_steps[user_id]["part1"] = message.content.strip()
            await message.reply("Blockers / additional note")
            return

        elif user_steps[user_id]["step"] == 1 and user_steps[user_id]["part1"] != "":
            user_steps[user_id]["part2"] = message.content.strip()
            session_state["morning"]["replied_users"].add(user_id)

            save_update(
                today_str,
                message.author,
                "morning",
                user_steps[user_id]["part1"],
                user_steps[user_id]["part2"],
                now_str
            )

            await message.reply("Great")
            await update_status_message(message.channel, guild, "morning")
            return

    # EVENING FLOW
    if session_state["evening"]["active"] and user_id not in session_state["evening"]["replied_users"]:
        user_steps = session_state["evening"]["user_steps"]

        if user_id not in user_steps:
            user_steps[user_id] = {
                "step": 1,
                "part1": "",
                "part2": ""
            }
            await message.reply("What did you achieve today?")
            return

        elif user_steps[user_id]["step"] == 1 and user_steps[user_id]["part1"] == "":
            user_steps[user_id]["part1"] = message.content.strip()
            await message.reply("Pending / blockers")
            return

        elif user_steps[user_id]["step"] == 1 and user_steps[user_id]["part1"] != "":
            user_steps[user_id]["part2"] = message.content.strip()
            session_state["evening"]["replied_users"].add(user_id)

            save_update(
                today_str,
                message.author,
                "evening",
                user_steps[user_id]["part1"],
                user_steps[user_id]["part2"],
                now_str
            )

            await message.reply("Great job")
            await update_status_message(message.channel, guild, "evening")
            return


bot.run(TOKEN)